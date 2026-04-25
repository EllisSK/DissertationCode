# SPDX-FileCopyrightText: 2026 Ellis Sinclair-Kent
#
# SPDX-License-Identifier: GPL-2.0-only
import warnings

import numpy as np
import pandas as pd

from openpyxl import load_workbook
from pathlib import Path
from tqdm import tqdm, trange
from scipy.stats import t

def scan_data_directory():
    data_dir = Path("data/") 

    contents = list(data_dir.iterdir())

    if not (data_dir / "BarrierExperiments.csv").exists():
        if (data_dir / "LabData.xlsx").exists():
            construct_lab_data_csv()
        else:
            # Download from archive
            pass
            
    if not (data_dir / "ManningsNExperiments.csv").exists():
        # Download from archive
        pass

def construct_lab_data_csv():
    path = Path("data/BarrierExperiments.csv")
    
    data = read_lab_data()

    data.to_csv(path, index=False)

def parse_float(val):
    if val is None or str(val).strip().upper() in ("N/A", "NA", ""):
        return None
    try:
        return float(val)
    except (ValueError, TypeError):
        return None

def read_lab_data() -> pd.DataFrame:
    path = Path("data/LabData.xlsx")

    column_names = [
        "Barrier Setup",
        "Operation Mode",
        "Set Flow (l/s)",
        "X Position (mm)",
        "Y Position (mm)",
        "Depth (mm)",
    ]

    book = load_workbook(filename=path, read_only=True, data_only=True)
    config_sheet_names = [s for s in book.sheetnames if s.startswith("Config ")]
    data_list = []

    print("Loading Lab Data. Progress:")
    for sheet_name in tqdm(config_sheet_names):
        sheet = book[sheet_name]

        barrier_setup = "{}-{}-{}".format(sheet["F2"].value, sheet["F3"].value, sheet["F4"].value)
        operation_mode = sheet["F5"].value
        set_flow = parse_float(sheet["B2"].value)

        for row in sheet["B11:D25"]:
            x_pos = parse_float(row[0].value)
            y_pos = parse_float(row[1].value)
            depth = parse_float(row[2].value)

            if x_pos is None or y_pos is None or depth is None:
                continue

            x_pos += 5000

            data_list.append(
                {
                    "Barrier Setup": barrier_setup,
                    "Operation Mode": operation_mode,
                    "Set Flow (l/s)": set_flow,
                    "X Position (mm)": x_pos,
                    "Y Position (mm)": y_pos,
                    "Depth (mm)": depth,
                }
            )

    return pd.DataFrame(data_list, columns=column_names)

def read_barrier_data() -> pd.DataFrame:
    path = Path("data/BarrierExperiments.csv")
    
    df = pd.read_csv(path)

    df = df[df["X Position (mm)"] < 5000]
    
    df_mean = df.groupby(
        ["Barrier Setup", "Operation Mode", "Set Flow (l/s)"], 
        as_index=False
    ).agg(
        **{"Mean Upstream Depth (mm)": ("Depth (mm)", "mean")}
    )
    
    return df_mean

def read_friction_data() -> pd.DataFrame:
    path = Path("data/ManningsNExperiments.csv")

    df = pd.read_csv(path)

    df = df.groupby(["Set Flow (l/s)", "Incline (%)", "X Position (mm)"])["Depth (mm)"].mean().reset_index()

    return df

def analyse_friction_data(data: pd.DataFrame):
    data["AOD (m)"] = ((10000 - data["X Position (mm)"]) / 1000) * (data["Incline (%)"] / 100)
    data["Velocity Head (m)"] = np.power((data["Set Flow (l/s)"] / data["Depth (mm)"]), 2) / (2 * 9.81)
    data["Depth (m)"] = data["Depth (mm)"] / 1000

    data["Total Head (m)"] = data["AOD (m)"] + data["Velocity Head (m)"] + data["Depth (m)"]

    data["X Position (m)"] = data["X Position (mm)"] / 1000

    def get_head_slope(group):
        slope, _ = np.polyfit(group["X Position (m)"], group["Total Head (m)"], 1)
        return slope
    
    slope_data = data.groupby(["Set Flow (l/s)", "Incline (%)"]).apply(get_head_slope).reset_index(name="Free Surface Slope")

    data = data.merge(slope_data, on=["Set Flow (l/s)", "Incline (%)"], how="left")
    data["Sf"] = -data["Free Surface Slope"]
    data["P (m)"] = 1 + (2 * data["Depth (m)"])

    data["Composite n"] = np.sqrt((data["Sf"] * np.power(data["Depth (m)"], 10/3)) / (np.power(data["Set Flow (l/s)"] / 1000, 2) * np.power(data["P (m)"], 4/3)))

    x_vals = (2 * data["Depth (m)"]) / 2
    y_vals = (data["P (m)"] * np.power(data["Composite n"], 1.5)) / 2
    slope, intercept = np.polyfit(x_vals, y_vals, 1)

    bed_n = np.power(intercept, 2/3)
    wall_n = np.power(slope, 2/3)

    return bed_n, wall_n

def write_friction_report(name: str, path: Path):
    fric = read_friction_data()
    
    data = analyse_friction_data(fric)

    df = pd.DataFrame([data], columns=["Bed", "Wall"])

    df.to_csv(path / name, index=False)

def run_monte_carlo_analysis(csv_data_path: Path, model, report_path: Path, trials: int = 200000):
    df = pd.read_csv(csv_data_path)
    
    unique_conditions = df[["Barrier Setup", "Operation Mode", "Set Flow (l/s)"]].drop_duplicates()
    condition_to_flow_samples = {}
    
    D = 0.350
    area = np.pi * (D / 2) ** 2
    velocity_error_m_s = 0.002
    flow_error_from_velocity_ls = area * velocity_error_m_s * 1000
    
    for _, row in unique_conditions.iterrows():
        nominal_flow = row["Set Flow (l/s)"]
        
        relative_error = 0.002 * nominal_flow
        total_flow_half_width = relative_error + flow_error_from_velocity_ls
        
        samples = np.random.uniform(
            nominal_flow - total_flow_half_width, 
            nominal_flow + total_flow_half_width, 
            size=trials
        )
        
        key = (row["Barrier Setup"], row["Operation Mode"], row["Set Flow (l/s)"])
        condition_to_flow_samples[key] = samples

    grouped = df.groupby(["Barrier Setup", "Operation Mode", "Set Flow (l/s)", "X Position (mm)"])
    
    n_groups = len(grouped)
    mc_depths = np.zeros((n_groups, trials))
    
    group_keys = []
    for i, (name, group) in enumerate(grouped):
        group_keys.append(name)
        n = len(group["Depth (mm)"])
        mean = group["Depth (mm)"].mean()
        std = group["Depth (mm)"].std(ddof=1)
        
        if n > 1 and std > 0:
            scale = std / np.sqrt(n)
            samples = t.rvs(df=n-1, loc=mean, scale=scale, size=trials)
        else:
            samples = np.full(trials, mean)
            
        mc_depths[i, :] = np.maximum(samples, 0)

    stats_results = {
        "RMSE": np.zeros(trials),
        "MAE": np.zeros(trials),
        "Absolute Bias": np.zeros(trials),
        "Variability Ratio": np.zeros(trials),
        "Correlation": np.zeros(trials),
        "KGE": np.zeros(trials),
        "R Squared": np.zeros(trials)
    }
    
    has_fit = hasattr(model, "fit") and hasattr(model, "optimal")
    opt_is_array = False
    num_coeffs = 1
    
    if has_fit:
        opt_is_array = isinstance(model.optimal, (list, np.ndarray))
        if opt_is_array:
            num_coeffs = len(model.optimal)
            for k in range(num_coeffs):
                stats_results[f"Optimised Coefficient {k+1}"] = np.zeros(trials)
        else:
            stats_results["Optimised Coefficient"] = np.zeros(trials)
            
        orig_df = model.df.copy() if hasattr(model, "df") else None
        orig_optimal = np.copy(model.optimal) if opt_is_array else model.optimal
        orig_popt = model.popt.copy() if hasattr(model, "popt") else None
        orig_pcov = model.pcov.copy() if hasattr(model, "pcov") else None

    with warnings.catch_warnings():
        warnings.simplefilter("ignore")
        
        for j in trange(trials, desc=f"MC: {model.name}"):
            trial_rows = []
            for i, name in enumerate(group_keys):
                condition_key = (name[0], name[1], name[2])
                trial_rows.append({
                    "Barrier Setup": name[0],
                    "Operation Mode": name[1],
                    "Set Flow (l/s)": condition_to_flow_samples[condition_key][j],
                    "X Position (mm)": name[3],
                    "Y Position (mm)": 0.0,
                    "Depth (mm)": mc_depths[i, j]
                })
            
            trial_df = pd.DataFrame(trial_rows)
            trial_df = trial_df[trial_df["X Position (mm)"] < 5000]
            
            trial_df = trial_df.groupby(
                ["Barrier Setup", "Operation Mode", "Set Flow (l/s)"], 
                as_index=False
            ).agg(
                **{"Mean Upstream Depth (mm)": ("Depth (mm)", "mean")}
            )
            
            processed_df = model._create_model_dataframe(trial_df)
            
            if has_fit:
                model.df = processed_df
                try:
                    model.fit()
                    if opt_is_array:
                        for k in range(num_coeffs):
                            stats_results[f"Optimised Coefficient {k+1}"][j] = model.optimal[k]
                    else:
                        stats_results["Optimised Coefficient"][j] = model.optimal
                except Exception:
                    if opt_is_array:
                        for k in range(num_coeffs):
                            stats_results[f"Optimised Coefficient {k+1}"][j] = np.nan
                    else:
                        stats_results["Optimised Coefficient"][j] = np.nan
            
            obj_metrics = model._calculate_objective_functions(processed_df)
            rmse, mae, bias, var, corr, kge, r_squared = obj_metrics[:7]
            
            stats_results["RMSE"][j] = rmse
            stats_results["MAE"][j] = mae
            stats_results["Absolute Bias"][j] = bias
            stats_results["Variability Ratio"][j] = var
            stats_results["Correlation"][j] = corr
            stats_results["KGE"][j] = kge
            stats_results["R Squared"][j] = r_squared

    if has_fit:
        if orig_df is not None:
            model.df = orig_df
        model.optimal = orig_optimal
        if orig_popt is not None:
            model.popt = orig_popt
        if orig_pcov is not None:
            model.pcov = orig_pcov

    def shortest_coverage_interval(data, alpha=0.95):
        data = np.sort(data[~np.isnan(data)])
        n_samples = len(data)
        
        if n_samples == 0:
            return np.nan, np.nan
            
        p = int(np.floor(alpha * n_samples))
        if p == 0:
            return data[0], data[-1]
        widths = data[p:] - data[:n_samples - p]
        min_idx = np.argmin(widths)
        return data[min_idx], data[min_idx + p]

    file_exists = report_path.exists()
    mode = "a" if file_exists else "w"
    
    with open(report_path, mode) as f:
        if file_exists:
            f.write("\n")
        f.write("Monte Carlo Analysis (95% CI Shortest Coverage Interval)\n")
            
        for stat_name, data in stats_results.items():
            lower, upper = shortest_coverage_interval(data, 0.95)
            f.write(f"{stat_name}: [{lower}, {upper}]\n")